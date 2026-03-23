import openpyxl
import json
import re
import sys

# ---- Team name normalization (from name_map.py) ----
TEAM_ALIASES = {
    "duke": "Duke", "duke blue devils": "Duke", "duke 1": "Duke",
    "michigan": "Michigan", "michigan wolverines": "Michigan", "michigan 1": "Michigan",
    "arizona": "Arizona", "arizona wildcats": "Arizona", "arizona 1": "Arizona",
    "florida": "Florida", "florida gators": "Florida", "florida 1": "Florida",
    "uconn": "UConn", "connecticut": "UConn", "uconn huskies": "UConn", "connecticut huskies": "UConn",
    "houston": "Houston", "houston cougars": "Houston", "houston 2": "Houston",
    "iowa state": "Iowa State", "iowa st": "Iowa State", "iowa st.": "Iowa State",
    "iowa state cyclones": "Iowa State", "iowa st. 2": "Iowa State", "iowa st 2": "Iowa State",
    "purdue": "Purdue", "purdue boilermakers": "Purdue", "purdue 2": "Purdue",
    "michigan state": "Michigan State", "michigan st": "Michigan State", "michigan st.": "Michigan State",
    "michigan state spartans": "Michigan State", "michigan st. 3": "Michigan State", "michigan st 3": "Michigan State",
    "illinois": "Illinois", "illinois fighting illini": "Illinois", "illinois 3": "Illinois",
    "gonzaga": "Gonzaga", "gonzaga bulldogs": "Gonzaga",
    "virginia": "Virginia", "virginia cavaliers": "Virginia",
    "alabama": "Alabama", "alabama crimson tide": "Alabama",
    "kansas": "Kansas", "kansas jayhawks": "Kansas",
    "nebraska": "Nebraska", "nebraska cornhuskers": "Nebraska",
    "arkansas": "Arkansas", "arkansas razorbacks": "Arkansas",
    "st. john's": "St. John's", "saint john's": "St. John's", "st john's": "St. John's",
    "saint johns": "St. John's", "st. john's red storm": "St. John's",
    "st john's (ny)": "St. John's", "saint john's (ny)": "St. John's",
    "st. john's (ny)": "St. John's", "st. johns": "St. John's",
    "texas tech": "Texas Tech", "texas tech red raiders": "Texas Tech",
    "vanderbilt": "Vanderbilt", "vanderbilt commodores": "Vanderbilt",
    "wisconsin": "Wisconsin", "wisconsin badgers": "Wisconsin",
    "louisville": "Louisville", "louisville cardinals": "Louisville",
    "north carolina": "North Carolina", "north carolina tar heels": "North Carolina", "unc": "North Carolina",
    "n. carolina": "North Carolina", "n carolina": "North Carolina",
    "tennessee": "Tennessee", "tennessee volunteers": "Tennessee",
    "byu": "BYU", "brigham young": "BYU", "byu cougars": "BYU", "brigham young cougars": "BYU",
    "kentucky": "Kentucky", "kentucky wildcats": "Kentucky",
    "saint mary's": "Saint Mary's", "st. mary's": "Saint Mary's", "st mary's": "Saint Mary's",
    "saint mary's college": "Saint Mary's", "st. mary's ca": "Saint Mary's",
    "st mary's ca": "Saint Mary's", "saint mary's ca": "Saint Mary's",
    "saint mary's gaels": "Saint Mary's", "st. mary's college": "Saint Mary's",
    "st mary's, cal.": "Saint Mary's",
    "saint mary's college gaels": "Saint Mary's", "saint mary's (ca)": "Saint Mary's",
    "saint marys": "Saint Mary's", "st marys": "Saint Mary's",
    "saint mary\u2019s": "Saint Mary's", "saint mary\u2019s college": "Saint Mary's",
    "saint mary\u2019s gaels": "Saint Mary's", "saint mary\u2019s (ca)": "Saint Mary's",
    "miami fl": "Miami FL", "miami (fl)": "Miami FL", "miami hurricanes": "Miami FL",
    "miami fl 7": "Miami FL", "miami": "Miami FL",
    "ucla": "UCLA", "ucla bruins": "UCLA", "california-los angeles": "UCLA",
    "clemson": "Clemson", "clemson tigers": "Clemson",
    "villanova": "Villanova", "villanova wildcats": "Villanova",
    "ohio state": "Ohio State", "ohio st": "Ohio State", "ohio st.": "Ohio State",
    "ohio state buckeyes": "Ohio State",
    "georgia": "Georgia", "georgia bulldogs": "Georgia",
    "tcu": "TCU", "texas christian": "TCU", "tcu horned frogs": "TCU",
    "iowa": "Iowa", "iowa hawkeyes": "Iowa",
    "saint louis": "Saint Louis", "st. louis": "Saint Louis", "st louis": "Saint Louis",
    "saint louis billikens": "Saint Louis",
    "utah state": "Utah State", "utah st": "Utah State", "utah st.": "Utah State",
    "utah state aggies": "Utah State",
    "ucf": "UCF", "central florida": "UCF", "ucf knights": "UCF",
    "missouri": "Missouri", "missouri tigers": "Missouri",
    "santa clara": "Santa Clara", "santa clara broncos": "Santa Clara",
    "texas a&m": "Texas A&M", "texas am": "Texas A&M", "texas a&m aggies": "Texas A&M",
    "south florida": "South Florida", "s. florida": "South Florida", "usf": "South Florida",
    "south florida bulls": "South Florida", "south florida 11": "South Florida",
    "s florida": "South Florida",
    "vcu": "VCU", "virginia commonwealth": "VCU", "vcu rams": "VCU",
    "va commonwealth": "VCU",
    "nc state": "NC State", "north carolina state": "NC State", "north carolina st.": "NC State",
    "n.c. state": "NC State", "nc state wolfpack": "NC State", "n carolina st.": "NC State",
    "n. carolina st.": "NC State",
    "texas": "Texas", "texas longhorns": "Texas",
    "smu": "SMU", "southern methodist": "SMU", "smu mustangs": "SMU",
    "southern meth.": "SMU", "southern methodist 11": "SMU",
    "miami oh": "Miami OH", "miami (oh)": "Miami OH", "miami ohio": "Miami OH",
    "miami (oh) redhawks": "Miami OH", "miami oh 11": "Miami OH",
    "miami redhawks": "Miami OH",
    "northern iowa": "Northern Iowa", "n. iowa": "Northern Iowa", "uni": "Northern Iowa",
    "northern iowa panthers": "Northern Iowa",
    "mcneese": "McNeese", "mcneese st": "McNeese", "mcneese st.": "McNeese",
    "mcneese state": "McNeese", "mcneese cowboys": "McNeese", "mcneese state cowboys": "McNeese",
    "high point": "High Point", "high point panthers": "High Point",
    "akron": "Akron", "akron zips": "Akron",
    "hofstra": "Hofstra", "hofstra pride": "Hofstra",
    "troy": "Troy", "troy trojans": "Troy", "troy st.": "Troy",
    "hawaii": "Hawaii", "hawai'i": "Hawaii", "hawaii rainbow warriors": "Hawaii",
    "hawai'i rainbow warriors": "Hawaii", "hawai\u02bbi": "Hawaii", "hawai\u2019i": "Hawaii",
    "penn": "Penn", "pennsylvania": "Penn", "pennsylvania quakers": "Penn", "penn quakers": "Penn",
    "north dakota state": "North Dakota State", "north dakota st": "North Dakota State",
    "n dakota st": "North Dakota State", "n. dakota st.": "North Dakota State",
    "north dakota state bison": "North Dakota State", "north dakota st.": "North Dakota State",
    "n dakota st.": "North Dakota State", "n. dakota st": "North Dakota State",
    "wright state": "Wright State", "wright st": "Wright State", "wright st.": "Wright State",
    "wright state raiders": "Wright State",
    "kennesaw state": "Kennesaw State", "kennesaw st": "Kennesaw State",
    "kennesaw": "Kennesaw State", "kennesaw state owls": "Kennesaw State",
    "kennesaw st.": "Kennesaw State",
    "furman": "Furman", "furman paladins": "Furman",
    "idaho": "Idaho", "idaho vandals": "Idaho",
    "tennessee state": "Tennessee State", "tennessee st": "Tennessee State",
    "tennessee st.": "Tennessee State", "tennessee state tigers": "Tennessee State",
    "tenn. st.": "Tennessee State", "tenn st": "Tennessee State", "tenn st.": "Tennessee State",
    "queens": "Queens", "queens university": "Queens", "queens (nc)": "Queens",
    "queens university royals": "Queens", "queens nc": "Queens",
    "queens (nc) royals": "Queens", "queens royals": "Queens", "queens univ": "Queens",
    "siena": "Siena", "siena saints": "Siena",
    "liu": "LIU", "long island": "LIU", "long island university": "LIU",
    "liu sharks": "LIU", "liu brooklyn": "LIU", "long island sharks": "LIU",
    "long island university sharks": "LIU",
    "howard": "Howard", "howard bison": "Howard", "howard u.": "Howard",
    "umbc": "UMBC", "maryland-baltimore co.": "UMBC", "md-baltimore": "UMBC",
    "umbc retrievers": "UMBC", "maryland baltimore co.": "UMBC", "maryland baltimore co": "UMBC",
    "lehigh": "Lehigh", "lehigh mountain hawks": "Lehigh",
    "prairie view a&m": "Prairie View A&M", "prairie view": "Prairie View A&M",
    "pv a&m": "Prairie View A&M", "prairie view a&m panthers": "Prairie View A&M",
    "cal baptist": "Cal Baptist", "california baptist": "Cal Baptist",
    "cal baptist lancers": "Cal Baptist", "california baptist lancers": "Cal Baptist",
    "cal baptist 13": "Cal Baptist", "ca baptist": "Cal Baptist", "calif baptist": "Cal Baptist",
    "north carolina state wolfpack": "NC State",
    "unc tar heels": "North Carolina",
    "c florida": "UCF", "n iowa": "Northern Iowa",
    "miami (fla.)": "Miami FL", "u miami (fl)": "Miami FL",
    "miami university (oh)": "Miami OH", "miami-ohio": "Miami OH",
    # Smart apostrophe variants
    "st. john\u2019s": "St. John's", "st john\u2019s": "St. John's", "saint john\u2019s": "St. John's",
    "st. mary\u2019s": "Saint Mary's", "st mary\u2019s": "Saint Mary's",
    "st. mary\u2019s ca": "Saint Mary's", "st mary\u2019s ca": "Saint Mary's",
    "st. marys": "Saint Mary's",
}

TOURNAMENT_TEAMS = [
    "Duke", "Michigan", "Arizona", "Florida", "Houston", "Iowa State", "Purdue", "Illinois",
    "Gonzaga", "UConn", "Michigan State", "St. John's", "Virginia", "Vanderbilt",
    "Arkansas", "Nebraska", "Tennessee", "Louisville", "North Carolina", "Wisconsin",
    "Alabama", "Kentucky", "Texas Tech", "UCLA", "Ohio State", "BYU", "Saint Mary's",
    "Georgia", "Clemson", "Saint Louis", "Villanova", "Santa Clara", "NC State", "TCU",
    "Texas A&M", "VCU", "South Florida", "SMU", "Texas", "Missouri", "UCF", "Akron",
    "McNeese", "Northern Iowa", "Miami OH", "High Point", "Hofstra", "Cal Baptist",
    "Hawaii", "North Dakota State", "Wright State", "Troy", "Kennesaw State", "Penn",
    "Idaho", "Tennessee State", "Queens", "Furman", "UMBC", "Siena", "Howard", "LIU",
    "Lehigh", "Prairie View A&M", "Iowa", "Utah State"
]

def normalize_team(name):
    if not name or str(name).strip() == '' or str(name) == 'None':
        return None
    name = str(name).strip()
    # Remove record in parens
    name = re.sub(r'\s*\([\d]+-[\d]+\)', '', name)
    # Remove newline content
    name = name.split('\n')[0].strip()
    # Remove emoji
    name = re.sub(r'[🏀📉🤕🔒✅▲▼]', '', name).strip()
    # Remove trailing seed numbers
    name = re.sub(r'\s+\d+$', '', name).strip()
    # Remove leading rank numbers like "1. Duke Blue Devils"
    name = re.sub(r'^\d+\.?\s*', '', name).strip()
    # Remove " seed," type text
    name = re.sub(r'\s*\d+\s*seed.*', '', name, flags=re.IGNORECASE).strip()
    
    clean = name.lower().strip()
    if clean in TEAM_ALIASES:
        return TEAM_ALIASES[clean]
    return None

def to_float(v):
    """Convert a value to float, handling embedded rank strings like '128.2\n4'"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', '-', '#ERROR!', 'None', '---'):
        return None
    # Split on newline, take first part
    s = s.split('\n')[0].strip()
    # Remove leading dot for values like ".9813"
    try:
        return float(s)
    except:
        return None

# Initialize result
result = {team: {} for team in TOURNAMENT_TEAMS}
summary = {}

BASE = "/Users/chrisdell/Downloads/drive-download-20260322T062153Z-1-001"

def extract_sheet(wb_path, sheet_name, team_col, value_cols, data_start, key_names, label):
    """
    Generic extractor.
    team_col: 1-indexed column for team name
    value_cols: list of 1-indexed columns for values
    data_start: first data row
    key_names: list of key names matching value_cols
    """
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    matched = 0
    unmatched = []
    for r in range(data_start, ws.max_row + 1):
        raw_name = ws.cell(r, team_col).value
        team = normalize_team(raw_name)
        if team and team in result:
            matched += 1
            for col, key in zip(value_cols, key_names):
                val = to_float(ws.cell(r, col).value)
                if val is not None:
                    result[team][key] = val
    wb.close()
    found_teams = set(t for t in TOURNAMENT_TEAMS if any(k in result[t] for k in key_names))
    missing = [t for t in TOURNAMENT_TEAMS if t not in found_teams]
    summary[label] = {"matched": len(found_teams), "missing": missing}
    return found_teams

# 1. KENPOM
print("Extracting KENPOM...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "KENPOM", 2, [5, 6, 8, 10], 3,
              ["kp_em", "kp_o", "kp_d", "kp_t"], "KENPOM")

# 2. BARTTORVIK
print("Extracting BARTTORVIK...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "BARTTORVIK", 2, [6, 7, 8, 23], 3,
              ["bt_oe", "bt_de", "bt_barthag", "bt_t"], "BARTTORVIK")

# 3. DUNKEL
print("Extracting DUNKEL...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "DUNKEL", 2, [4], 2,
              ["dunkel"], "DUNKEL")

# 4. DEEPMETRICS
print("Extracting DEEPMETRICS...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "DEEPMETRICS", 2, [7, 8, 9], 2,
              ["deep_net", "deep_o", "deep_d"], "DEEPMETRICS")

# 5. DRATINGS
print("Extracting DRATINGS...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "DRATINGS", 1, [2], 2,
              ["drat"], "DRATINGS")

# 6. VERSUS
print("Extracting VERSUS...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "VERSUS", 2, [3], 2,
              ["versus"], "VERSUS")

# 7. ELO
print("Extracting ELO...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "ELO", 1, [3], 2,
              ["elo"], "ELO")

# 8. COLLEY
print("Extracting COLLEY...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "COLLEY", 2, [4], 3,
              ["colley"], "COLLEY")

# 9. TEAMRANK
print("Extracting TEAMRANK...")
extract_sheet(f"{BASE}/CPI TABLES.xlsx", "TEAMRANK", 3, [2], 2,
              ["tmrk"], "TEAMRANK")

# 10. INCC
print("Extracting INCC...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "INCC", 2, [7], 6,
              ["incc"], "INCC")

# 11. WAYWARD
print("Extracting WAYWARD...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "WAYWARD", 2, [6], 2,
              ["waywd"], "WAYWARD")

# 12. WHEELUS
print("Extracting WHEELUS...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "WHEELUS", 2, [4], 2,
              ["wheel"], "WHEELUS")

# 13. WHITLOCK
print("Extracting WHITLOCK...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "WHITLOCK", 1, [2], 2,
              ["whtlk"], "WHITLOCK")

# 14. JAMESENG
print("Extracting JAMESENG...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "JAMESENG", 2, [4], 2,
              ["jeng"], "JAMESENG")

# 15. TAILSMAN
print("Extracting TAILSMAN...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "TAILSMAN", 2, [5], 4,
              ["talis"], "TAILSMAN")

# 16. QRI
print("Extracting QRI...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "QRI", 3, [1], 3,
              ["qri"], "QRI")

# 17. WHO EARNED IT
print("Extracting WHO EARNED IT...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "WHO EARNED IT", 2, [4], 4,
              ["whoe"], "WHO EARNED IT")

# 18. AMSTS
print("Extracting AMSTS...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "AMSTS", 2, [5], 2,
              ["amsts"], "AMSTS")

# 19. OMNI
print("Extracting OMNI...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "OMNI", 2, [3], 7,
              ["omni"], "OMNI")

# 20. JNG
print("Extracting JNG...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "JNG", 2, [3], 3,
              ["jng"], "JNG")

# 21. SRCBB
print("Extracting SRCBB...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "SRCBB", 2, [16, 14, 15, 17, 18, 19], 14,
              ["srcbb_srs", "srcbb_osrs", "srcbb_dsrs", "srcbb_ortg", "srcbb_drtg", "srcbb_nrtg"], "SRCBB")

# 22. AEI
print("Extracting AEI...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "AEI", 2, [3], 2,
              ["aei"], "AEI")

# 23. RAMS - col 8 seems to be main rating (looks like a composite score)
print("Extracting RAMS...")
extract_sheet(f"{BASE}/CPI 2ND BATCH.xlsx", "RAMS", 2, [8], 6,
              ["rams"], "RAMS")

# 24. SIMMONS
print("Extracting SIMMONS...")
extract_sheet(f"{BASE}/CPI 3RD BATCH.xlsx", "SIMMONS", 2, [5], 2,
              ["simm"], "SIMMONS")

# 25. TPR
print("Extracting TPR...")
extract_sheet(f"{BASE}/CPI 3RD BATCH.xlsx", "TPR", 2, [3], 1,
              ["tpr"], "TPR")

# 26. PACKARD
print("Extracting PACKARD...")
extract_sheet(f"{BASE}/CPI 3RD BATCH.xlsx", "PACKARD", 2, [3], 2,
              ["pack"], "PACKARD")

# 27. SWISH
print("Extracting SWISH...")
extract_sheet(f"{BASE}/SWISH.xlsx", "Sheet1", 1, [2, 5, 7, 9], 2,
              ["sw_rating", "sw_tempo", "sw_oeff", "sw_deff"], "SWISH")

# 28. HASLA OFFENSE
print("Extracting HASLA OFFENSE...")
extract_sheet(f"{BASE}/HASLA.xlsx", "OFFENSE", 2, [3], 4,
              ["has_oeff"], "HASLA_OFFENSE")

# 29. HASLA DEFENSE
print("Extracting HASLA DEFENSE...")
extract_sheet(f"{BASE}/HASLA.xlsx", "DEFENSE", 2, [3], 6,
              ["has_deff"], "HASLA_DEFENSE")

# ---- Fill missing keys with null ----
ALL_KEYS = [
    "kp_em", "kp_o", "kp_d", "kp_t",
    "bt_oe", "bt_de", "bt_barthag", "bt_t",
    "sw_rating", "sw_tempo", "sw_oeff", "sw_deff",
    "has_oeff", "has_deff",
    "incc", "deep_net", "deep_o", "deep_d", "dunkel",
    "elo", "aei", "colley",
    "srcbb_srs", "srcbb_osrs", "srcbb_dsrs", "srcbb_ortg", "srcbb_drtg", "srcbb_nrtg",
    "tmrk", "versus",
    "waywd", "wheel", "whtlk",
    "jeng", "talis", "qri",
    "whoe", "amsts", "omni", "jng",
    "simm", "tpr", "pack",
    "drat", "rams"
]

for team in TOURNAMENT_TEAMS:
    for key in ALL_KEYS:
        if key not in result[team]:
            result[team][key] = None

# Write JSON
out_path = "/Users/chrisdell/Downloads/bracketx-project/raw_ratings.json"
with open(out_path, 'w') as f:
    json.dump(result, f, indent=2)

print(f"\nWrote {out_path}")
print(f"\n{'='*60}")
print("EXTRACTION SUMMARY")
print(f"{'='*60}")
for sys_name, info in sorted(summary.items()):
    m = info["matched"]
    miss = info["missing"]
    status = f"{m}/66" if m < 66 else f"{m}/66 ALL"
    # Some systems won't have all 66 (small schools)
    print(f"  {sys_name:20s}: {status}")
    if miss and len(miss) <= 15:
        print(f"    Missing: {', '.join(miss)}")
    elif miss:
        print(f"    Missing {len(miss)} teams")

# Check which teams have the most/fewest ratings
print(f"\n{'='*60}")
print("TEAM COVERAGE")
print(f"{'='*60}")
for team in TOURNAMENT_TEAMS:
    filled = sum(1 for k in ALL_KEYS if result[team][k] is not None)
    if filled < 20:
        print(f"  {team:25s}: {filled}/{len(ALL_KEYS)} ratings")

