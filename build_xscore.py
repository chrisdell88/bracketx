#!/usr/bin/env python3
"""
BracketX X-Score Composite Builder
Phase 1: Extract ratings from all 37+ systems, Z-score normalize, weight, compute X-Score
"""

import openpyxl
import csv
import re
import json
import numpy as np
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════
# STEP 0: Define the 68 tournament teams (from verified bracket)
# ═══════════════════════════════════════════════════════════════

TOURNAMENT_TEAMS_RAW = [
    # (standard_name, seed, region, conference, record, aliases)
    ("Duke", 1, "East", "ACC", "32-2", ["Duke", "Duke Blue Devils"]),
    ("Michigan", 1, "Midwest", "Big Ten", "31-3", ["Michigan", "Michigan Wolverines"]),
    ("Arizona", 1, "West", "Big 12", "32-2", ["Arizona", "Arizona Wildcats"]),
    ("Florida", 1, "South", "SEC", "26-7", ["Florida", "Florida Gators"]),
    ("UConn", 2, "East", "Big East", "29-5", ["UConn", "Connecticut", "Connecticut Huskies", "UConn Huskies"]),
    ("Houston", 2, "South", "Big 12", "28-6", ["Houston", "Houston Cougars"]),
    ("Iowa State", 2, "Midwest", "Big 12", "27-7", ["Iowa State", "Iowa St", "Iowa St.", "Iowa State Cyclones"]),
    ("Purdue", 2, "West", "Big Ten", "27-8", ["Purdue", "Purdue Boilermakers"]),
    ("Michigan State", 3, "East", "Big Ten", "25-7", ["Michigan State", "Michigan St", "Michigan St.", "Michigan State Spartans"]),
    ("Illinois", 3, "South", "Big Ten", "24-8", ["Illinois", "Illinois Fighting Illini"]),
    ("Gonzaga", 3, "West", "WCC", "30-3", ["Gonzaga", "Gonzaga Bulldogs"]),
    ("Virginia", 3, "Midwest", "ACC", "29-5", ["Virginia", "Virginia Cavaliers"]),
    ("Alabama", 4, "Midwest", "SEC", "23-9", ["Alabama", "Alabama Crimson Tide"]),
    ("Kansas", 4, "East", "Big 12", "23-10", ["Kansas", "Kansas Jayhawks"]),
    ("Nebraska", 4, "South", "Big Ten", "26-6", ["Nebraska", "Nebraska Cornhuskers"]),
    ("Arkansas", 4, "West", "SEC", "26-8", ["Arkansas", "Arkansas Razorbacks"]),
    ("St. John's", 5, "East", "Big East", "28-6", ["St. John's", "Saint John's", "St John's", "Saint Johns", "St. John's Red Storm", "St John's (NY)", "Saint John's (NY)"]),
    ("Texas Tech", 5, "Midwest", "Big 12", "22-10", ["Texas Tech", "Texas Tech Red Raiders"]),
    ("Vanderbilt", 5, "South", "SEC", "26-8", ["Vanderbilt", "Vanderbilt Commodores"]),
    ("Wisconsin", 5, "West", "Big Ten", "24-10", ["Wisconsin", "Wisconsin Badgers"]),
    ("Louisville", 6, "East", "ACC", "23-10", ["Louisville", "Louisville Cardinals"]),
    ("North Carolina", 6, "South", "ACC", "24-8", ["North Carolina", "N. Carolina", "UNC", "North Carolina Tar Heels"]),
    ("Tennessee", 6, "Midwest", "SEC", "22-11", ["Tennessee", "Tennessee Volunteers"]),
    ("BYU", 6, "West", "Big 12", "23-11", ["BYU", "Brigham Young", "Brigham Young Cougars", "BYU Cougars"]),
    ("Kentucky", 7, "Midwest", "SEC", "21-13", ["Kentucky", "Kentucky Wildcats"]),
    ("Saint Mary's", 7, "South", "WCC", "27-5", ["Saint Mary's", "St. Mary's", "St Mary's", "Saint Mary's College", "St. Mary's CA", "Saint Mary's Gaels", "St Mary's CA"]),
    ("Miami FL", 7, "West", "ACC", "25-8", ["Miami FL", "Miami (FL)", "Miami Hurricanes", "Miami"]),
    ("UCLA", 7, "East", "Big Ten", "23-11", ["UCLA", "UCLA Bruins", "California-Los Angeles"]),
    ("Clemson", 8, "South", "ACC", "24-10", ["Clemson", "Clemson Tigers"]),
    ("Villanova", 8, "West", "Big East", "24-8", ["Villanova", "Villanova Wildcats"]),
    ("Ohio State", 8, "East", "Big Ten", "21-12", ["Ohio State", "Ohio St", "Ohio St.", "Ohio State Buckeyes"]),
    ("Georgia", 8, "Midwest", "SEC", "22-10", ["Georgia", "Georgia Bulldogs"]),
    ("TCU", 9, "East", "Big 12", "22-11", ["TCU", "Texas Christian", "TCU Horned Frogs"]),
    ("Iowa", 9, "South", "Big Ten", "21-12", ["Iowa", "Iowa Hawkeyes"]),
    ("Saint Louis", 9, "Midwest", "A-10", "28-5", ["Saint Louis", "St. Louis", "St Louis", "Saint Louis Billikens"]),
    ("Utah State", 9, "West", "MWC", "28-6", ["Utah State", "Utah St", "Utah St.", "Utah State Aggies"]),
    ("UCF", 10, "East", "Big 12", "21-11", ["UCF", "Central Florida", "UCF Knights"]),
    ("Missouri", 10, "West", "SEC", "20-12", ["Missouri", "Missouri Tigers"]),
    ("Santa Clara", 10, "Midwest", "WCC", "26-8", ["Santa Clara", "Santa Clara Broncos"]),
    ("Texas A&M", 10, "South", "SEC", "21-11", ["Texas A&M", "Texas AM", "Texas A&M Aggies"]),
    ("South Florida", 11, "East", "AAC", "25-8", ["South Florida", "S. Florida", "USF", "South Florida Bulls"]),
    ("VCU", 11, "South", "A-10", "27-7", ["VCU", "Virginia Commonwealth", "VA Commonwealth", "VCU Rams"]),
    ("NC State", 11, "West", "ACC", "20-13", ["NC State", "North Carolina State", "North Carolina St.", "N.C. State", "NC State Wolfpack"]),
    ("Texas", 11, "West", "SEC", "18-14", ["Texas", "Texas Longhorns"]),  # First Four
    ("SMU", 11, "Midwest", "ACC", "20-13", ["SMU", "Southern Methodist", "Southern Meth.", "SMU Mustangs"]),
    ("Miami OH", 11, "Midwest", "MAC", "31-1", ["Miami OH", "Miami (OH)", "Miami Ohio", "Miami (OH) RedHawks"]),  # First Four
    ("Northern Iowa", 12, "East", "MVC", "23-12", ["Northern Iowa", "N. Iowa", "UNI", "Northern Iowa Panthers"]),
    ("McNeese", 12, "South", "Southland", "28-5", ["McNeese", "McNeese St", "McNeese St.", "McNeese State", "McNeese Cowboys"]),
    ("High Point", 12, "West", "Big South", "30-4", ["High Point", "High Point Panthers"]),
    ("Akron", 12, "Midwest", "MAC", "29-5", ["Akron", "Akron Zips"]),
    ("Hofstra", 13, "Midwest", "CAA", "24-10", ["Hofstra", "Hofstra Pride"]),
    ("Troy", 13, "South", "Sun Belt", "22-11", ["Troy", "Troy Trojans", "Troy St."]),
    ("Hawaii", 13, "West", "Big West", "24-8", ["Hawaii", "Hawai'i", "Hawaii Rainbow Warriors", "Hawai'i Rainbow Warriors"]),
    ("Penn", 14, "South", "Ivy", "18-11", ["Penn", "Pennsylvania", "Pennsylvania Quakers"]),
    ("North Dakota State", 14, "East", "Summit", "27-7", ["North Dakota State", "North Dakota St", "N Dakota St", "N. Dakota St.", "North Dakota State Bison", "North Dakota St."]),
    ("Wright State", 14, "Midwest", "Horizon", "23-11", ["Wright State", "Wright St", "Wright St."]),
    ("Kennesaw State", 14, "West", "CUSA", "21-13", ["Kennesaw State", "Kennesaw St", "Kennesaw", "Kennesaw State Owls"]),
    ("Furman", 15, "East", "SoCon", "22-12", ["Furman", "Furman Paladins"]),
    ("Idaho", 15, "South", "Big Sky", "21-14", ["Idaho", "Idaho Vandals"]),
    ("Tennessee State", 15, "Midwest", "OVC", "23-9", ["Tennessee State", "Tennessee St", "Tennessee St.", "Tennessee State Tigers"]),
    ("Queens", 15, "West", "ASUN", "21-13", ["Queens", "Queens University", "Queens University Royals"]),
    ("Siena", 16, "East", "MAAC", "23-11", ["Siena", "Siena Saints"]),
    ("LIU", 16, "West", "NEC", "24-10", ["LIU", "Long Island", "Long Island University", "Long Island University Sharks"]),
    ("Howard", 16, "Midwest", "MEAC", "23-10", ["Howard", "Howard Bison"]),  # First Four
    ("UMBC", 16, "Midwest", "Am. East", "24-8", ["UMBC", "Maryland-Baltimore Co.", "MD-Baltimore", "UMBC Retrievers"]),  # First Four
    ("Lehigh", 16, "South", "Patriot", "18-16", ["Lehigh", "Lehigh Mountain Hawks"]),  # First Four
    ("Prairie View A&M", 16, "South", "SWAC", "20-14", ["Prairie View A&M", "Prairie View", "PV A&M", "Prairie View A&M Panthers"]),  # First Four
]

# Build alias -> standard name map
ALIAS_MAP = {}
for team_info in TOURNAMENT_TEAMS_RAW:
    standard = team_info[0]
    aliases = team_info[5]
    for alias in aliases:
        ALIAS_MAP[alias.lower().strip()] = standard

# Also add some common variations
EXTRA_ALIASES = {
    "michigan st.": "Michigan State",
    "michigan st": "Michigan State", 
    "iowa st": "Iowa State",
    "iowa st.": "Iowa State",
    "ohio st": "Ohio State",
    "ohio st.": "Ohio State",
    "utah st": "Utah State",
    "utah st.": "Utah State",
    "north dakota st": "North Dakota State",
    "north dakota st.": "North Dakota State",
    "n dakota st": "North Dakota State",
    "wright st": "Wright State",
    "wright st.": "Wright State",
    "kennesaw st": "Kennesaw State",
    "kennesaw st.": "Kennesaw State",
    "tennessee st": "Tennessee State",
    "tennessee st.": "Tennessee State",
    "mcneese st": "McNeese",
    "mcneese st.": "McNeese",
    "connecticut": "UConn",
    "brigham young": "BYU",
    "virginia commonwealth": "VCU",
    "st john's": "St. John's",
    "saint john's": "St. John's",
    "st. john's (ny)": "St. John's",
    "saint johns": "St. John's",
    "saint john's (ny)": "St. John's",
    "st. mary's college": "Saint Mary's",
    "st mary's ca": "Saint Mary's",
    "st. mary's ca": "Saint Mary's",
    "st. mary's": "Saint Mary's",
    "st mary's": "Saint Mary's",
    "saint mary's ca": "Saint Mary's",
    "central florida": "UCF",
    "southern methodist": "SMU",
    "southern meth.": "SMU",
    "north carolina state": "NC State",
    "north carolina st.": "NC State",
    "n. carolina st.": "NC State",
    "south florida": "South Florida",
    "s. florida": "South Florida",
    "miami (fl)": "Miami FL",
    "miami hurricanes": "Miami FL",
    "miami (oh)": "Miami OH",
    "miami ohio": "Miami OH",
    "n. iowa": "Northern Iowa",
    "northern iowa": "Northern Iowa",
    "uni": "Northern Iowa",
    "texas christian": "TCU",
    "texas a&m": "Texas A&M",
    "texas am": "Texas A&M",
    "st. louis": "Saint Louis",
    "st louis": "Saint Louis",
    "california-los angeles": "UCLA",
    "long island": "LIU",
    "md-baltimore": "UMBC",
    "maryland-baltimore co.": "UMBC",
    "prairie view": "Prairie View A&M",
    "pv a&m": "Prairie View A&M",
    "hawai'i": "Hawaii",
    "penn": "Penn",
    "pennsylvania": "Penn",
    "n carolina st.": "NC State",
    "duke 1": "Duke",
    "arizona 1": "Arizona",
    "michigan 1": "Michigan",
    "florida 1": "Florida",
    "houston 2": "Houston",
}

for k, v in EXTRA_ALIASES.items():
    ALIAS_MAP[k.lower().strip()] = v


def normalize_team(name):
    """Convert any team name variant to our standard name"""
    if not name:
        return None
    # Clean up
    name = str(name).strip()
    # Remove record in parens like "Duke (32-2)"
    name = re.sub(r'\s*\([\d]+-[\d]+\)', '', name)
    # Remove seed numbers like "Duke 1" or trailing digits
    name = re.sub(r'\s+\d+$', '', name)
    # Remove emoji and special chars
    name = re.sub(r'[🏀📉🤕🔒]', '', name)
    # Remove newline content like "Duke\n 1 seed, ✅"
    name = name.split('\n')[0].strip()
    
    clean = name.lower().strip()
    
    if clean in ALIAS_MAP:
        return ALIAS_MAP[clean]
    
    # Try partial matches
    for alias, standard in ALIAS_MAP.items():
        if clean == alias or alias.startswith(clean) or clean.startswith(alias):
            return standard
    
    return None  # Not a tournament team


# ═══════════════════════════════════════════════════════════════
# STEP 1: Define tier weights
# ═══════════════════════════════════════════════════════════════

TIER1_WEIGHT = 3.0  # Tempo-adjusted per-possession
TIER2_WEIGHT = 1.0  # Score-margin, no confirmed tempo
TIER3_WEIGHT = 0.5  # Simplest models

# System definitions: (name, tier, source_description)
SYSTEMS = {}  # Will be populated as we extract data


# ═══════════════════════════════════════════════════════════════
# STEP 2: Extract ratings from all sources
# ═══════════════════════════════════════════════════════════════

# Master ratings dict: {standard_team_name: {system_name: raw_rating}}
ratings = defaultdict(dict)

print("=" * 60)
print("EXTRACTING RATINGS FROM ALL SOURCES")
print("=" * 60)

# --- CPI_TABLES.xlsx (Batch 1) ---
wb1 = openpyxl.load_workbook('/mnt/user-data/uploads/CPI_TABLES.xlsx')

# KENPOM - Tier 1
print("\n[KENPOM] - Tier 1")
ws = wb1['KENPOM']
count = 0
for r in range(3, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    netrtg_raw = str(ws.cell(r, 5).value or "")
    team = normalize_team(team_raw)
    if team and netrtg_raw:
        try:
            val = float(netrtg_raw.replace('=+', '').replace('+', ''))
            ratings[team]['kenpom'] = val
            count += 1
        except:
            pass
SYSTEMS['kenpom'] = ('KenPom', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# BARTTORVIK - Tier 1
print("\n[BARTTORVIK] - Tier 1")
ws = wb1['BARTTORVIK']
count = 0
for r in range(3, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    barthag_raw = str(ws.cell(r, 8).value or "")
    team = normalize_team(team_raw)
    if team and barthag_raw:
        try:
            val = float(barthag_raw.split('\n')[0].replace('.', '', 1).replace('.', '.'))
            # Barthag is like .9813 - convert to float
            val_clean = barthag_raw.split('\n')[0].strip()
            val = float(val_clean)
            ratings[team]['barttorvik'] = val
            count += 1
        except:
            pass
SYSTEMS['barttorvik'] = ('BartTorvik', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# DUNKEL - Tier 2
print("\n[DUNKEL] - Tier 2")
ws = wb1['DUNKEL']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 4).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['dunkel'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['dunkel'] = ('Dunkel', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# TEAMRANK - Tier 2
print("\n[TEAMRANK] - Tier 2")
ws = wb1['TEAMRANK']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 3).value or "")
    rating_raw = ws.cell(r, 2).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['teamrankings'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['teamrankings'] = ('TeamRankings', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# COLLEY - Tier 2
print("\n[COLLEY] - Tier 2")
ws = wb1['COLLEY']
count = 0
for r in range(3, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 4).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['colley'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['colley'] = ('Colley', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# DEEPMETRICS - Tier 2
print("\n[DEEPMETRICS] - Tier 2")
ws = wb1['DEEPMETRICS']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    net_raw = ws.cell(r, 7).value
    team = normalize_team(team_raw)
    if team and net_raw:
        try:
            ratings[team]['deepmetric'] = float(net_raw)
            count += 1
        except:
            pass
SYSTEMS['deepmetric'] = ('Deep Metric', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# DRATINGS - Tier 2
print("\n[DRATINGS] - Tier 2")
ws = wb1['DRATINGS']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 1).value or "")
    # Extract team name from "1. Duke Blue Devils (32-2)"
    m = re.match(r'\d+\.\s*(.+?)\s*\(', team_raw)
    if m:
        team_name = m.group(1).strip()
    else:
        team_name = team_raw
    rating_raw = ws.cell(r, 2).value
    team = normalize_team(team_name)
    if team and rating_raw:
        try:
            ratings[team]['dratings'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['dratings'] = ('DRatings', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# VERSUS - Tier 2
print("\n[VERSUS] - Tier 2")
ws = wb1['VERSUS']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "").split('\n')[0].strip()
    rating_raw = ws.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['versus'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['versus'] = ('Versus', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# ELO (Warren Nolan) - Tier 3
print("\n[ELO] - Tier 3")
ws = wb1['ELO']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 1).value or "")
    elo_raw = ws.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and elo_raw:
        try:
            ratings[team]['elo_wn'] = float(elo_raw)
            count += 1
        except:
            pass
SYSTEMS['elo_wn'] = ('ELO (Warren Nolan)', TIER3_WEIGHT)
print(f"  Extracted {count} teams")

# --- Evan Miya (data.csv) - Tier 1 ---
print("\n[EVAN MIYA] - Tier 1")
count = 0
with open('/mnt/user-data/uploads/data.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        team = normalize_team(row.get('team', ''))
        bpr = row.get('bpr', '')
        if team and bpr:
            try:
                ratings[team]['evanmiya'] = float(bpr)
                count += 1
            except:
                pass
SYSTEMS['evanmiya'] = ('Evan Miya', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# --- COOPER Elo (CjvhY.csv) - Tier 1 ---
print("\n[COOPER ELO] - Tier 1")
count = 0
with open('/mnt/user-data/uploads/CjvhY.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        team = normalize_team(row.get('sb_name', ''))
        elo = row.get('xelo_n', '')
        if team and elo:
            try:
                ratings[team]['cooper_elo'] = float(elo)
                count += 1
            except:
                pass
SYSTEMS['cooper_elo'] = ('COOPER Elo', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# --- SwishStats (SWISH.xlsx) - Tier 1 ---
print("\n[SWISHSTATS] - Tier 1")
count = 0
wb_swish = openpyxl.load_workbook('/mnt/user-data/uploads/SWISH.xlsx')
ws = wb_swish['Sheet1']
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 1).value or "")
    rating_raw = ws.cell(r, 2).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['swishstats'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['swishstats'] = ('SwishStats', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# --- Haslametrics (HASLA__3_.xlsx) - Tier 1 ---
print("\n[HASLAMETRICS] - Tier 1")
count = 0
wb_hasla = openpyxl.load_workbook('/mnt/user-data/uploads/HASLA__3_.xlsx')
ws_off = wb_hasla['OFFENSE']
ws_def = wb_hasla['DEFENSE']
# Build offense dict
off_eff = {}
for r in range(4, ws_off.max_row + 1):
    team_raw = str(ws_off.cell(r, 2).value or "")
    eff_raw = ws_off.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and eff_raw:
        try:
            off_eff[team] = float(eff_raw)
        except:
            pass
# Build defense dict
def_eff = {}
for r in range(6, ws_def.max_row + 1):
    team_raw = str(ws_def.cell(r, 2).value or "")
    eff_raw = ws_def.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and eff_raw:
        try:
            def_eff[team] = float(eff_raw)
        except:
            pass
# Compute net efficiency
for team in off_eff:
    if team in def_eff:
        ratings[team]['haslametrics'] = off_eff[team] - def_eff[team]
        count += 1
SYSTEMS['haslametrics'] = ('Haslametrics', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# --- Massey own ratings (export__2_.csv) - Tier 2 ---
print("\n[MASSEY] - Tier 2")
count = 0
with open('/mnt/user-data/uploads/export__2_.csv', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    next(reader)  # skip header
    for row in reader:
        if len(row) >= 7:
            team = normalize_team(row[0])
            pwr = row[6]  # Power rating
            if team and pwr:
                try:
                    ratings[team]['massey'] = float(pwr)
                    count += 1
                except:
                    pass
SYSTEMS['massey'] = ('Massey', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# --- Power Rank (web scraped earlier) - Tier 1 ---
# This was scraped in conversation. Using data from thepowerrank.com
SYSTEMS['powerrank'] = ('Power Rank', TIER1_WEIGHT)

# --- INCC Stats - Tier 1 ---
print("\n[INCC] - Tier 1")
count = 0
wb2 = openpyxl.load_workbook('/mnt/user-data/uploads/CPI_2ND_BATCH.xlsx')
ws = wb2['INCC']
for r in range(6, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = str(ws.cell(r, 7).value or "")
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            val = float(rating_raw.replace('=+', '').replace('+', ''))
            ratings[team]['incc'] = val
            count += 1
        except:
            pass
SYSTEMS['incc'] = ('INCC Stats', TIER1_WEIGHT)
print(f"  Extracted {count} teams")

# --- Batch 2 systems from CPI_2ND_BATCH.xlsx ---

# AEI - Tier 3
print("\n[AEI] - Tier 3")
ws = wb2['AEI']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    elo_raw = ws.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and elo_raw:
        try:
            ratings[team]['aei'] = float(elo_raw)
            count += 1
        except:
            pass
SYSTEMS['aei'] = ('AEI', TIER3_WEIGHT)
print(f"  Extracted {count} teams")

# WAYWARD/BWE - Tier 2
print("\n[WAYWARD] - Tier 2")
ws = wb2['WAYWARD']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 6).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['wayward'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['wayward'] = ('Wayward/BWE', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# WHEELUS - Tier 2
print("\n[WHEELUS] - Tier 2")
ws = wb2['WHEELUS']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 4).value  # Final Rating
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['wheelus'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['wheelus'] = ('Wheelus', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# WHITLOCK - Tier 2
print("\n[WHITLOCK] - Tier 2")
ws = wb2['WHITLOCK']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 1).value or "")
    rating_raw = ws.cell(r, 2).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['whitlock'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['whitlock'] = ('Whitlock', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# JAMESENG - Tier 2
print("\n[JAMESENG] - Tier 2")
ws = wb2['JAMESENG']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 4).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['jameseng'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['jameseng'] = ('JamesEng', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# TALISMAN RED - Tier 2
print("\n[TALISMAN] - Tier 2")
ws = wb2['TAILSMAN']
count = 0
for r in range(4, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 5).value  # RATING column
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['talisman'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['talisman'] = ('Talisman Red', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# QRI - Tier 2
print("\n[QRI] - Tier 2")
ws = wb2['QRI']
count = 0
for r in range(3, ws.max_row + 1):
    team_raw = str(ws.cell(r, 3).value or "")
    rating_raw = ws.cell(r, 1).value  # QRI Points
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['qri'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['qri'] = ('QRI', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# WHO EARNED IT - Tier 2
print("\n[WHOEARNEDIT] - Tier 2")
ws = wb2['WHO EARNED IT']
count = 0
for r in range(4, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 4).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['whoearnedit'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['whoearnedit'] = ('WhoEarnedIt', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# OMNI - Tier 2
print("\n[OMNI] - Tier 2")
ws = wb2['OMNI']
count = 0
for r in range(7, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['omni'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['omni'] = ('Omni Rankings', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# JNG - Tier 2
print("\n[JNG] - Tier 2")
ws = wb2['JNG']
count = 0
for r in range(3, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 3).value
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['jng'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['jng'] = ('JNG (HoopsHD)', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# AMSTS - Tier 2
print("\n[AMSTS] - Tier 2")
ws = wb2['AMSTS']
count = 0
for r in range(2, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    rating_raw = ws.cell(r, 5).value  # Overall Rating
    team = normalize_team(team_raw)
    if team and rating_raw:
        try:
            ratings[team]['amsts'] = float(rating_raw)
            count += 1
        except:
            pass
SYSTEMS['amsts'] = ('AllMySportsTeams', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# RAMS - Tier 2
print("\n[RAMS] - Tier 2")
ws = wb2['RAMS']
count = 0
for r in range(5, ws.max_row + 1):
    team_raw = str(ws.cell(r, 2).value or "")
    # Use Marg (margin) column as primary
    marg_raw = ws.cell(r, 6).value
    team = normalize_team(team_raw)
    if team and marg_raw:
        try:
            val = str(marg_raw).replace('+', '')
            ratings[team]['rams'] = float(val)
            count += 1
        except:
            pass
SYSTEMS['rams'] = ('RAMS', TIER2_WEIGHT)
print(f"  Extracted {count} teams")

# Simmons (from web scrape earlier - stored in thepowerrank data)
SYSTEMS['simmons'] = ('Simmons', TIER2_WEIGHT)
SYSTEMS['sonnymoore'] = ('Sonny Moore', TIER2_WEIGHT)
SYSTEMS['bcmoore'] = ('BC Moore', TIER2_WEIGHT)

# Wolfe - Tier 3
SYSTEMS['wolfe'] = ('Wolfe', TIER3_WEIGHT)


# ═══════════════════════════════════════════════════════════════
# STEP 3: Z-Score Normalize within 68-team field
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("Z-SCORE NORMALIZATION")
print("=" * 60)

tournament_teams = [t[0] for t in TOURNAMENT_TEAMS_RAW]

# For each system, compute mean and std across tournament teams
z_scores = defaultdict(dict)
system_coverage = {}

for sys_name in SYSTEMS:
    # Get all tournament team ratings for this system
    vals = []
    teams_with_data = []
    for team in tournament_teams:
        if sys_name in ratings.get(team, {}):
            vals.append(ratings[team][sys_name])
            teams_with_data.append(team)
    
    if len(vals) < 10:
        print(f"  [{sys_name}] Only {len(vals)} tournament teams - SKIPPING")
        system_coverage[sys_name] = len(vals)
        continue
    
    mean = np.mean(vals)
    std = np.std(vals)
    system_coverage[sys_name] = len(vals)
    
    if std == 0:
        print(f"  [{sys_name}] Zero variance - SKIPPING")
        continue
    
    for team in tournament_teams:
        if sys_name in ratings.get(team, {}):
            z = (ratings[team][sys_name] - mean) / std
            z_scores[team][sys_name] = z
    
    print(f"  [{sys_name}] {len(vals)}/68 teams | mean={mean:.3f} std={std:.3f}")


# ═══════════════════════════════════════════════════════════════
# STEP 4: Compute weighted X-Score
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("COMPUTING X-SCORES")
print("=" * 60)

x_scores = {}

for team in tournament_teams:
    weighted_sum = 0
    total_weight = 0
    
    for sys_name, (sys_label, tier_weight) in SYSTEMS.items():
        if sys_name in z_scores.get(team, {}):
            z = z_scores[team][sys_name]
            weighted_sum += z * tier_weight
            total_weight += tier_weight
    
    if total_weight > 0:
        x_scores[team] = weighted_sum / total_weight
    else:
        x_scores[team] = 0

# Sort by X-Score
ranked = sorted(x_scores.items(), key=lambda x: x[1], reverse=True)

# Get team info
team_info = {}
for t in TOURNAMENT_TEAMS_RAW:
    team_info[t[0]] = {'seed': t[1], 'region': t[2], 'conf': t[3], 'record': t[4]}


print("\n" + "=" * 60)
print("BRACKETX — X-SCORE COMPOSITE RANKINGS")
print(f"Systems used: {sum(1 for s in system_coverage if system_coverage[s] >= 10)}")
print(f"Tier 1 (3x): {sum(1 for s,v in SYSTEMS.items() if v[1]==3.0 and system_coverage.get(s,0)>=10)}")
print(f"Tier 2 (1x): {sum(1 for s,v in SYSTEMS.items() if v[1]==1.0 and system_coverage.get(s,0)>=10)}")
print(f"Tier 3 (0.5x): {sum(1 for s,v in SYSTEMS.items() if v[1]==0.5 and system_coverage.get(s,0)>=10)}")
print("=" * 60)

print(f"\n{'Rk':>3} {'Team':<22} {'Seed':>4} {'Region':<10} {'Record':<7} {'X-Score':>8} {'Systems':>7}")
print("-" * 70)

for i, (team, score) in enumerate(ranked):
    info = team_info[team]
    num_systems = len(z_scores.get(team, {}))
    print(f"{i+1:>3} {team:<22} {info['seed']:>4} {info['region']:<10} {info['record']:<7} {score:>+8.4f} {num_systems:>7}")

# Save to JSON for dashboard
output = {
    'rankings': [],
    'systems': {k: {'name': v[0], 'weight': v[1], 'coverage': system_coverage.get(k, 0)} 
                for k, v in SYSTEMS.items() if system_coverage.get(k, 0) >= 10},
    'methodology': {
        'tier1_weight': TIER1_WEIGHT,
        'tier2_weight': TIER2_WEIGHT, 
        'tier3_weight': TIER3_WEIGHT,
        'normalization': 'Z-score within 68-team tournament field',
    }
}

for i, (team, score) in enumerate(ranked):
    info = team_info[team]
    entry = {
        'rank': i + 1,
        'team': team,
        'seed': info['seed'],
        'region': info['region'],
        'conf': info['conf'],
        'record': info['record'],
        'xscore': round(score, 4),
        'num_systems': len(z_scores.get(team, {})),
        'z_scores': {k: round(v, 4) for k, v in z_scores.get(team, {}).items()},
    }
    output['rankings'].append(entry)

with open('/home/claude/xscore_output.json', 'w') as f:
    json.dump(output, f, indent=2)

print(f"\n\nResults saved to /home/claude/xscore_output.json")
